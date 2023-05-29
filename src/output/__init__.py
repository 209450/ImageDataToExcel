from openpyxl import load_workbook


def append_excel_file(output_file_name, sheet_name, data):
    wb = load_workbook(output_file_name)

    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name)
        sheet.title = sheet_name

    sheet.append(data)
    wb.save(output_file_name)
