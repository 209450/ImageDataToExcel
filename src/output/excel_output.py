from openpyxl import load_workbook


def create_new_sheet(work_book, sheet_name):
    sheet = work_book.create_sheet(sheet_name)
    sheet.title = sheet_name
    return sheet


def get_current_sheet(work_book, sheet_name):
    if sheet_name in work_book.sheetnames:
        sheet = work_book[sheet_name]
    else:
        sheet = create_new_sheet(work_book, sheet_name)
    return sheet
