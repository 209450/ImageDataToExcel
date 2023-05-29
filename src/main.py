import argparse
import json
import sys
import os
import easyocr
import numpy as np
from PIL import Image

from PyQt5.QtWidgets import QApplication, QMessageBox
from openpyxl import Workbook, load_workbook

from data_structures.file_type import check_input_file_type
from gui.change_rectangle_coordinates_dialog import FormChangeRectangleCoordinates
from gui.image_with_rectangles_window import ImageWindowWithRectangles
from gui.scanned_data_check_dialog import ScannedDataCheckDialog
from ocr.read_text_from_image import read_text_from_image_rectangles
from output.excel_output import get_current_sheet, create_new_sheet


def open_change_rectangle_window(image_path, input_table_rectangles):
    image_window = ImageWindowWithRectangles(image_path)
    image_window.show()

    placement_of_rectangles_is_not_correct = True
    while placement_of_rectangles_is_not_correct:

        image_window.clear_drawings()
        for input_table_index, rectangle in enumerate(input_table_rectangles):
            image_window.draw_rectangle(rectangle.top_left, rectangle.bottom_right)

            label_padding_y = 5
            label_coordinates = (rectangle.top_left[0], rectangle.top_left[1] - label_padding_y)
            image_window.draw_label(label_coordinates, str(input_table_index))

        message_box_answer = QMessageBox.question(image_window, "Rectangles placement",
                                                  "Do the coordinates of rectangle are "
                                                  "valid?")
        if message_box_answer == QMessageBox.Yes:
            placement_of_rectangles_is_not_correct = False
        else:
            form = FormChangeRectangleCoordinates(image_window, input_table_rectangles)
            form.show()

            form_result = form.exec_()
            if form_result:
                input_table_rectangles = list(form.get_fields_values())

    image_window.deleteLater()
    return input_table_rectangles


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", action="store", dest="input",
                        help="path to image file", required=True)
    parser.add_argument("--output", action="store", dest="output",
                        help="path to output excel file", required=True)

    return parser.parse_args()


if __name__ == '__main__':
    print("Program started")
    print(f"current dir: {os.getcwd()}")

    args = parse_args()
    input_file_path = args.input
    output_file_path = args.output
    print(args.input)

    file_type = check_input_file_type(input_file_path)
    print(file_type)

    try:
        table_rectangles = list(file_type.value[1])
    except IndexError:
        print(f"{file_type} can not be processed")
        print("Program ended")
        sys.exit(1)

    config_path = "config.json"
    with open(config_path, encoding="utf-8") as file:
        loaded_json = json.load(file)
        output_data = loaded_json[file_type.value[0]].copy()

    input_image = Image.open(input_file_path)

    output_tables = {}
    output_data_is_not_correct = True
    app_image_window = QApplication(sys.argv)
    while output_data_is_not_correct:

        table_rectangles = open_change_rectangle_window(input_file_path, table_rectangles)
        app_image_window.closeAllWindows()

        tables_names = file_type.value[2]
        output_tables_names = dict.fromkeys(tables_names, "")
        output_tables_fields = file_type.value[2]
        for table_name, rectangle in zip(output_tables_names, table_rectangles):
            output_tables[table_name] = read_text_from_image_rectangles(output_tables_fields[table_name], input_image,
                                                                        rectangle)

        dialog_results = []
        for table_name, rectangle in zip(output_tables.keys(), table_rectangles):
            dialog = ScannedDataCheckDialog(input_file_path, rectangle, output_tables[table_name])
            dialog.show()

            dialog_result = dialog.exec_()
            dialog_results.append(dialog_result)
            if dialog_result:
                output_tables[table_name] = dialog.get_fields_values()

        if 0 in dialog_results:
            output_data_is_not_correct = True
        else:
            output_data_is_not_correct = False

    # for key, value in output_tables.items():
    #     print(f"{key}:{value}")

    cols_labels = ["file name"]
    for table_cols in output_tables.values():
        cols = list(table_cols.keys())
        cols_labels.extend(cols)

    output = [input_file_path]
    for table_cols in output_tables.values():
        output.extend(table_cols.values())

    if not os.path.exists(output_file_path):
        work_book = Workbook()
        work_book.save(output_file_path)
    else:
        work_book = load_workbook(output_file_path)

    sheet_name = file_type.value[0]
    if sheet_name not in work_book.sheetnames:
        current_sheet = create_new_sheet(work_book, sheet_name)
        current_sheet.append(cols_labels)
    else:
        current_sheet = get_current_sheet(work_book, sheet_name)
    current_sheet.append(output)

    work_book.save(output_file_path)
